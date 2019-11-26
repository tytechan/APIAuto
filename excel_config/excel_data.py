# encoding = utf - 8
# 用于获取excel中对应案例数据

from . import *
from preinfo_config.global_config import *
from openpyxl.styles import colors,fonts
import sys
import json

'''
【函数统计】
1、装饰器函数：jumpOrNot;
2、excel取值函数：getInterfaceData、makeJsonData（调用装饰器）、makeProcessData、getDataTransFormed；
3、excel写值函数：writeTextResult、loadProcessValue、clearProcessValue；
4、变量处理函数：latestReport、pickProcessDataRow、combineJson、checkTheMessage、notChooseNull、re_search、pinyinTransform.
'''

# ********************************* 装饰器及变量初始化函数 *********************************

def jumpOrNot(func):
    # 闭包中外函数中的变量指向的引用不可变
    numA = 0
    numB = 0
    def call_func(*args,**kwargs):
        nonlocal numA,numB
        result = func(*args,**kwargs)
        # print("args.__len__():",args.__len__())
        # print("kwargs:",kwargs)
        # print("result:",result)
        try:
            assert result == ""
            numA += 1
        except Exception as e:
            numB += 1
        finally:
            # print("空值数：",numA,"；正常取值数：",numB)

            if len(kwargs) > 0:
                # 判断是否继续执行
                if numB == 0:
                    result = "中断"
                # 重置计数
                numA = 0
                numB = 0

            return result
    return call_func


# ********************************* excel取值函数 *********************************

def getInterfaceData(strName):
    '''获取excel中对应案例sheet数据（包含每个案例初始化的操作）
    :param moduleName: 当前流程名称
    :param caseName: 当前案例名称
    :param strName: 待取字段名称
    :return:
    '''
    moduleName = get_value('MODULENAME')
    caseName = get_value('CASENAME')

    mySheet = excelObj.getSheetByName(moduleName)
    caseColumn = excelObj.getColumn(mySheet,Step_caseRowNum)

    # 流程中第一个案例在获取接口数据阶段，须进行之前执行结果的初始化
    if "test01" in caseName:
        clearProcessValue()

    # 获取对应案例行（流程sheet）
    for Looptime, case in enumerate(caseColumn[2:]):
        if case.value == caseName:
            caseRow = Looptime + 3
            break

    # 判断各字段对应列下是否有值，并取值
    myColumn = 1
    for myBox in excelObj.getRow(mySheet, 2):
        if myBox.value == None:
            break
        elif myBox.value == strName:
            myValue = excelObj.getCellOfValue(mySheet,rowNo=caseRow,colsNo=myColumn)
            print("🔼",strName,"的值为：",myValue)
            return str(myValue)
        myColumn += 1


def getDataTransFormed(*args,strName=None,whetherToInitialize=None):
    '''获取excel中“数据表”对应案例行的数据（针对页面字段和报文字段不一致情况）
    :param args:报文内容（“RESULT”）中对应路径
    :param strName:页面字段名称
    :param whetherToInitialize:中止判断节点
    :return:报文中对应字段
    '''
    myInfo = get_value("RESULT")
    myStr = makeJsonData(strName,whetherToInitialize=whetherToInitialize)
    myDict = combineJson(myInfo,*args)
    myDict = eval(myDict)

    for key in myDict.keys():
        if myDict[key] == myStr:
            print("🔎 转换后 '%s' 对应的值为：%s" %(myStr,key))
            return key

    print("❌ 未找到 %s 可转换的值！" %myStr)


@jumpOrNot
def makeJsonData(strName,whetherToInitialize=None):
    '''获取excel中“数据表”对应案例行的数据
    :param strName:待取字段名称
    :return:待取字段值
    '''
    dataSheetName = get_value("DATASHEETNAME")
    myRow = get_value('TESTROW')

    DataSheet = excelObj.getSheetByName(dataSheetName)

    # 判断各字段对应列下是否有值，并取值
    myColumn = 1
    for myBox in excelObj.getRow(DataSheet, 2):
        if myBox.value == None:
            break
        elif str(myBox.value) == strName:
            myValue = excelObj.getCellOfValue(DataSheet,rowNo=myRow,colsNo=myColumn)
            print("🔼 数据表中",strName,"的值为：",myValue)
            if myValue is None:
                return ""
            else:
                return str(myValue)
        myColumn += 1


def makeProcessData(strName,multiRow=None):
    '''获取excel中“数据表”对应案例行的数据，适用于取返回值
    :param strName:待取字段名称
    :return:待取字段值
    '''
    dataSheetName = get_value("DATASHEETNAME")
    DataSheet = excelObj.getSheetByName(dataSheetName)

    # 若该流程为非并发，则“数据表”sheet中数据行数固定；若为并发流程，行数为“multiRow”，从外部传入
    if multiRow:
        myRow = multiRow
    else:
        myRow = get_value('TESTROW')


    # 判断各字段对应列下是否有值，并取值
    myColumn = 1
    for myBox in excelObj.getRow(DataSheet, 2):
        if myBox.value == None:
            break
        elif myBox.value == strName:
            myValue = excelObj.getCellOfValue(DataSheet,rowNo=myRow,colsNo=myColumn)
            if strName == "#流程开关":
                if str(myValue) and str(myValue) != "":
                    print("📌流程开关中有值，流程终止！")
                else:
                    print("📌该案例正常执行！")
            else:
                print("🔼 数据表中",strName,"的值为：",myValue)

            if myValue is None:
                return ""
            else:
                return str(myValue)
        myColumn += 1


# ********************************* excel写值函数 *********************************

def writeTextResult(myRow=None):
    '''写入执行结果
    :param caseName:本案例名称
    :param result:成功/跳过/失败
    :param isFinal: 通过该参数控制，1）案例执行结果；2）流程执行结果
    :return:
    '''
    colorDict = {"成功":"green",
                 "跳过":"green",
                 "失败":"red",
                 "":None}

    dataDict = {"成功":"已使用",
                "跳过":"已使用",
                "失败":"未使用"}

    # 案例汇总页
    mainSheet = excelObj.getSheetByIndex(0)
    mainCaseColumn = excelObj.getColumn(mainSheet, Intro_caseRowNum)
    caseName = get_value('CASENAME')
    result = get_value('TESTRESULT')
    # 数据表sheet
    dataSheetName = get_value("DATASHEETNAME")
    DataSheet = excelObj.getSheetByName(dataSheetName)

    # 获取对应案例行
    for LooptimeNext, mainCase in enumerate(mainCaseColumn[1:]):
        if mainCase.value == caseName:
            mainCaseRow = LooptimeNext + 2
            break
    try:
        # print("案例写值时，result为：",result)
        excelObj.writeCell(mainSheet,result,rowNo=mainCaseRow,colsNo=Intro_testResult,style=colorDict[result])
        excelObj.writeCellCurrentTime(mainSheet,rowNo=mainCaseRow,colsNo=Intro_currentTimeResult)
    except Exception as e:
        print("********** “案例汇总”sheet 写入执行结果失败！ **********")
        raise e

    if myRow:
        # 数据表页
        # DataSheet = excelObj.getSheetByName("数据表")
        try:
            reportName = latestReport()
            excelObj.writeCell(DataSheet,reportName,rowNo=myRow,colsNo=Data_reportFileName)
            excelObj.writeCell(DataSheet,dataDict[result],rowNo=myRow,colsNo=Data_dataToBeUsed)
            excelObj.writeCell(DataSheet,result,rowNo=myRow,colsNo=Data_testResult,style=colorDict[result])
            excelObj.writeCellCurrentTime(DataSheet,rowNo=myRow,colsNo=Data_currentTimeResult)

        except Exception as e:
            print("********** “数据表”sheet 写入执行结果失败！ **********")
            raise e


def loadProcessValue(whichValue,*args,realValue=None):
    '''保存本案例返回值至“数据表”sheet中
    :param whichValue: 待保存字段名称
    :param realValue: 直接写入该值
    :param args:待保存字段在报文中路径
    '''

    dataSheetName = get_value("DATASHEETNAME")
    DataSheet = excelObj.getSheetByName(dataSheetName)

    if realValue is not None:
        valueReturned = realValue
    else:
        result = get_value('RESULT')
        valueReturned = combineJson(result,*args).replace("'","\"")

    # DataSheet = excelObj.getSheetByName("数据表")
    loopTime = get_value('TESTLOOPTIME')
    myRow = get_value('TESTROW')
    # RGBDict = {'red': 'FFFF3030', 'green': 'FF008B00'}

    myColumn = 1
    for myBox in excelObj.getRow(DataSheet, 2):
        if myBox.value == None:
            break
        elif myBox.value == whichValue:
            DataSheet.cell(row=myRow, column=myColumn).value = valueReturned

            if whichValue == "#流程开关":
                if valueReturned == "":
                    print("\n⏰ 开始执行新流程，流程数：",loopTime)
                else:
                    print("🎈流程即将结束！")

                    if valueReturned == "流程失败":
                        DataSheet.cell(row=myRow, column=myColumn).font = Font(color=colors.RED)
                    else:
                        DataSheet.cell(row=myRow, column=myColumn).font = Font(color=colors.BLACK)
            else:
                print("📦 返回值 %s 为: %s" %(whichValue.replace("#",""),valueReturned))

            return
        myColumn += 1


def clearProcessValue():
    '''
    清空在“数据表”sheet对应行中的返回值及之前的执行记录
    '''
    dataSheetName = get_value("DATASHEETNAME")
    myRow = get_value('TESTROW')
    DataSheet = excelObj.getSheetByName(dataSheetName)

    DataSheet.cell(row=myRow, column=Data_testResult).value = ""
    DataSheet.cell(row=myRow, column=Data_currentTimeResult).value = ""
    DataSheet.cell(row=myRow, column=Data_reportFileName).value = ""

    myColumn = 1
    for myBox in excelObj.getRow(DataSheet, 2):
        if myBox.value == None:
            break
        elif str(myBox.value).startswith("#"):
            DataSheet.cell(row=myRow, column=myColumn).value = ""
        myColumn += 1


# ********************************* 变量处理函数 *********************************

def latestReport():
    '''获取当前流程报告的文件名'''
    parentDirPath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    reportPath = parentDirPath + u"\\report"
    lists = os.listdir(reportPath)                                     #列出目录的下所有文件和文件夹保存到lists
    lists.sort(key=lambda fn:os.path.getmtime(reportPath + "\\" + fn)) #按时间排序
    file_new = lists[-1]      #获取最新的文件保存到file_new
    return file_new


def pickProcessDataRow(firstRow=None):
    '''定位该流程数据所在行
    :return: 本次流程调用数据所在行
    '''
    dataSheetName = get_value("DATASHEETNAME")
    DataSheet = excelObj.getSheetByName(dataSheetName)
    dataIsExecuteColumn = excelObj.getColumn(DataSheet, Data_dataIsExecute)

    if firstRow:
        originalRow = firstRow
    else:
        originalRow = 2

    noDataAccount = 0
    for Looptime, ExcuteMsg in enumerate(dataIsExecuteColumn[originalRow:]):
        if ExcuteMsg.value == None:
            if noDataAccount == 0:
                continue
            elif noDataAccount == 1:
                print("⛔ 已无足够可用数据，数据用完后执行终止!")
                os._exit(0)

        if firstRow:
            if ExcuteMsg.value != "已使用":
                myRow = originalRow + 1
                return myRow
            else:
                originalRow += 1
        else:
            if ExcuteMsg.value != "已使用":
                myRow = Looptime + 3
                return myRow

    # dataIsExecuteColumn对应当列有值的区域，当该区域循环结束后，直接结束进程
    print("⛔ 已无足够可用数据，数据用完后执行终止!")
    os._exit(0)


def combineJson(myMessage,*args):
    '''指定路径，定位报文中对应字段的值
    :param myMessage: 原始返回报文
    :param args:待处理字段路径
    '''
    # import json
    # myMessage = json.dumps(myMessage)
    myMessage = str(myMessage)
    try:
        # print("args.__len__():",args.__len__())
        for i in range(0,args.__len__()):
            if isinstance(args[i],int):
                myValue = str(args[i])
            else:
                myValue = '\"' + args[i] + '\"'
            myMessage += '[' + myValue + ']'
        finalMessage = str(eval(myMessage))
        return finalMessage
    except KeyError as e:
        print("😭 响应报文有误，请检查请求信息！ \n     %s \n" %myMessage)
        raise e


def checkTheMessage(*args,varNameInExcel=None,realValue=None):
    '''校验报文中具体字段值
    :param varNameInExcel:步骤sheet中待校验的字段名
    :param realValue:待校验预期值
    （若varValue存在，则 预期值 = varValue；
      若varValue不存在，则 预期值 = 根据varNameInExcel从excel获取的值）
    :param args:待处理字段在实际报文中的路径
    '''
    result = get_value('RESULT')

    try:
        if realValue:
            textToTest = realValue
        else:
            textToTest = getInterfaceData(varNameInExcel)

        myMessage = combineJson(result,*args)
        # print("myMessage:",myMessage)
        # print("textToTest:",textToTest)
        assert myMessage == textToTest, \
            "😭 响应报文中校验字段 '%s' 中， \n预期值 '%s' 和实际值 '%s' 不相等！" \
            %(args[args.__len__()-1],textToTest,myMessage)
        print("😜 响应报文中校验字段 '%s' 中， \n预期值 '%s' 和实际值 '%s' 相等！" \
            %(args[args.__len__()-1],textToTest,myMessage))
    except AssertionError as e:
        raise AssertionError(e)
    except Exception as e:
        raise e


def notChooseNull(varA,varB):
    '''若varA为“”，则返回varB，反之返回varA
    :param varA:一般对应excel中的值
    :param varB:一般对应脚本中的值
    :return:
    '''
    if varA != "":
        valuReturned = varA
    else:
        valuReturned = varB
    return valuReturned


def re_search(dict_data, key): #TODO：完善根据键值获取json中路径的方法
    for i in dict_data.keys():
        if i == key:
            return dict_data
        elif isinstance(dict_data[i], dict):
            tmp = re_search(dict_data[i], key)
            if tmp == None or tmp == {}:
                continue
            else:
                return tmp


def pinyinTransform(myStr):
    '''将汉字转换成拼音'''
    try:
        import pypinyin
        from pypinyin import pinyin, lazy_pinyin
        strTransformed = ''.join(lazy_pinyin(myStr))
        return strTransformed
    except Exception as e:
        raise e